import numpy as np
from elbotto.bots.training.manual_testing import is_none, get_model, create_test_matrix
# from elbotto.bots.training.trumpf_converter import trumpf_converter
# from elbotto.bots.training.trumpf_training import create_input
from openpyxl import load_workbook, Workbook


# def create_test_matrix(hand_cards, pushed_from_partner=False): #extract method
#     if isinstance(pushed_from_partner, bool) and pushed_from_partner:
#         shift = trumpf_converter(6)
#     else:
#         shift = trumpf_converter(0)
#
#     hand_card_list = fillup_card_list(hand_cards)
#
#     return create_input(hand_card_list, shift)


def manuel_test_list_input_predict(import_list_file, model=None):
    t_model = get_model(model)

    wb_input = load_workbook(import_list_file)
    sheets = wb_input.sheetnames
    sheet = wb_input[sheets[0]]
    input_sheet_matrix = tuple(sheet.rows)

    hand_cards_matrix = extract_lines_from_sheet(input_sheet_matrix)

    print('handcards: {}'.format(hand_cards_matrix))
    card_matrix = shrink_to_handcards(hand_cards_matrix)

    prediction_matrix = prepared_prediction_matrix(card_matrix)

    results = calc_prediction_matrix(prediction_matrix, t_model)

    safe_as_excelfile(card_matrix, results)


def calc_prediction_matrix(prediction_matrix, t_model):
    results = []
    for predict_input in prediction_matrix:
        result = t_model.predict(np.asarray(predict_input))
        result_list = matrix_to_list(result)
        print(
            'The prediction is: \n'
            ' hearts: {} \n diamonds: {} \n clubs: {} \n spades: {} \n OBEABE: {} \n UNDEUFE: {} \n SCHIEBE: {}'
                .format(result[0][0], result[0][1], result[0][2], result[0][3],
                        result[0][4], result[0][5], result[0][6]))
        results.append(result_list)
    return results


def prepared_prediction_matrix(card_matrix):
    prediction_matrix = []
    for cards_in_matrix in card_matrix:
        input_matrix = create_test_matrix(cards_in_matrix[0:9], cards_in_matrix[9])
        prediction_matrix.append(input_matrix)
    return prediction_matrix


def shrink_to_handcards(hand_cards_matrix):
    card_matrix = []
    for start_cards in hand_cards_matrix:
        if is_none(start_cards[0]):
            card_matrix.append(start_cards)
            print('Output the input_matrix: {}'.format(card_matrix))
        else:
            print("It's a None-Entry!!!")
    return card_matrix


def extract_lines_from_sheet(input_sheet_matrix):
    hand_cards_matrix = []
    for row_number in range(1, len(input_sheet_matrix)):
        row = input_sheet_matrix[row_number]
        input_list = []
        for cell in row:
            input_list.append(cell.value)

        print('InputListe: {}'.format(input_list))

        if is_none(input_list[len(input_list) - 1]):
            input_list[len(input_list) - 1] = True
        else:
            input_list[len(input_list) - 1] = False

        hand_cards_matrix.append(input_list)
    return hand_cards_matrix


def matrix_to_list(result):
    new_list = []
    for r in result[0]:
        new_list.append(r)
    print('New List: {}'.format(new_list))
    return new_list


def safe_as_excelfile(hand_cards, results, filename="./data/trumpf_cards/result_Trumpfwahl.xlsx"):
    length_of_output_matrix = len(hand_cards[0]) + len(results[0])

    wb = Workbook()
    ws = wb.active
    ws.title = "results trumpf prediction"

    create_table_title(length_of_output_matrix, ws)

    fillup_table(hand_cards, length_of_output_matrix, results, ws)

    wb.save(filename)


def fillup_table(hand_cards, length_of_output_matrix, results, ws):
    matrix_line_index = 0
    for row in ws.iter_rows(min_row=2, max_col=length_of_output_matrix, max_row=len(results) + 1):
        i = 0
        for cell in row:
            if i < len(hand_cards[0]):
                print(hand_cards[matrix_line_index][i])
                cell.value = hand_cards[matrix_line_index][i]
            else:
                print('{}'.format(results[matrix_line_index][i - len(hand_cards[0])]))
                cell.value = results[matrix_line_index][i - len(hand_cards[0])]
            i += 1
        matrix_line_index += 1


def create_table_title(length_of_output_matrix, ws):
    for first_row in ws.iter_rows(min_row=1, max_col=length_of_output_matrix, max_row=1):
        i = 0
        cell_names = ['Karte1', 'Karte2', 'Karte3', 'Karte4', 'Karte5', 'Karte6', 'Karte7', 'Karte8', 'Karte9',
                      'Geschoben?',
                      'Hearts', 'Diamonds', 'Clubs', 'Spades', 'Obeabe', 'Undeufe', 'Schieben']
        for cell in first_row:
            cell.value = cell_names[i]
            i += 1


# Input your hand cards into a excel list as the template excel sheet
# Set also in the list template in the column 'Geschoben?' or 'pushed?' a 'x' if your partner moved the trump decision to you.
# Set with model that model you want to test.
if __name__ == '__main__':
    manuel_test_list_input_predict(import_list_file="./data/trumpf_list.xlsx",
                                   model="./config/trumpf_network_model_final__2018-05-09_101936.h5")