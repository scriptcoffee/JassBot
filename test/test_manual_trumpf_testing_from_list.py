import pytest
from openpyxl import load_workbook, Workbook
from elbotto.bots.training.manual_trumpf_testing_from_list import matrix_to_list, \
                                                                    calc_prediction_matrix, \
                                                                    shrink_lines_to_handcards, \
                                                                    prepared_prediction_matrix, \
                                                                    extract_lines_from_sheet, \
    manuel_test_list_input_predict, create_table_title, safe_as_excelfile, fillup_table


@pytest.mark.parametrize("test_input, expected", [
    ("", None),
    (None, None),
    (['A1', 'A2', 'A3', 'A4', 'A5', 'A6'], None),
    (('A1', 'A2', 'A3', 'A4', 'A5', 'A6'), None),
    ((('A1', 'A2', 'A3', 'A4', 'A5'), ('B1', 'B2', 'B3', 'B4', 'B5'), ('C1', 'C2', 'C3', 'C4', 'C5')), None)
])
@pytest.mark.statistical
def test_extract_lines_from_sheet_wrong_inputs(test_input, expected):
    assert extract_lines_from_sheet(test_input) == expected


@pytest.mark.statistical
def test_extract_lines_from_sheet_empty_list():
    wb_input = load_workbook("./test/trumpf_test/test_trumpf_emptylist.xlsx")
    sheets = wb_input.sheetnames
    sheet = wb_input[sheets[0]]
    input_sheet_matrix = tuple(sheet.rows)

    result = extract_lines_from_sheet(input_sheet_matrix)

    assert result == []


@pytest.mark.statistical
def test_extract_lines_from_sheet_correct_inputs():
    expected = [['H6', 'C8', 'H9', 'HJ', 'D8', 'D9', 'D10', 'S8', 'S10', True],
                [None, None, None, None, None, None, None, None, None, False],
                ['H6', 'C8', 'H9', 'HJ', 'D8', 'D9', 'D10', 'S8', 'SA', False]]

    wb_input = load_workbook("./test/trumpf_test/test_trumpf_list.xlsx")
    sheets = wb_input.sheetnames
    sheet = wb_input[sheets[0]]
    input_sheet_matrix = tuple(sheet.rows)

    result = extract_lines_from_sheet(input_sheet_matrix)

    assert result == expected


@pytest.mark.parametrize("test_input, expected", [
    ([['H6', 'C8', 'H9', 'HJ', 'D8', 'D9', 'D10', 'S8', 'S10', True],
      [None, None, None, None, None, None, None, None, None, False],
      ['H6', 'C8', 'H9', 'HJ', 'D8', 'D9', 'D10', 'S8', 'SA', False]],
     [['H6', 'C8', 'H9', 'HJ', 'D8', 'D9', 'D10', 'S8', 'S10', True],
      ['H6', 'C8', 'H9', 'HJ', 'D8', 'D9', 'D10', 'S8', 'SA', False]]),
    ([['H9', 'H10', 'SJ', 'CA'], [None, 'D6', 'SK', 'HQ']],
     [['H9', 'H10', 'SJ', 'CA']]),
    ([['H9', 'H10', 'SJ', 'CA'], []],
     [['H9', 'H10', 'SJ', 'CA']]),
    ([['D6', 'SK', 'HQ']], [['D6', 'SK', 'HQ']]),
    (['D6', 'SK', 'HQ'], ['D6', 'SK', 'HQ']),
    ([None], []),
    (None, []),
    ([], []),
    ('test', []),
    (12, [])
])
@pytest.mark.statistical
def test_shrink_lines_to_handcards(test_input, expected):
    assert shrink_lines_to_handcards(test_input) == expected
